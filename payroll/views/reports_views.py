from django.http import HttpResponse
from openpyxl import Workbook
from rest_framework import status
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated, IsAdminUser
from django.db.models import Q, Sum, Avg, Count
from django.core.paginator import Paginator
from hr_management.models.hr_management_models import Employee
from payroll.models import Payroll, PayRun
from payroll.models.statutory_challan import StatutoryChallan
from payroll.utils.company_context import get_company_from_request
from payroll.utils.date_utils import get_financial_year
from rest_framework.parsers import MultiPartParser, FormParser
from django.utils.timezone import now
from openpyxl.styles import Font
from datetime import timedelta

class PTReportView(APIView):
    permission_classes = (IsAuthenticated,)

    def post(self, request):
        try:
            month = request.data.get("month")
            year = request.data.get("year")

            company = get_company_from_request(request)
            if not company:
                return Response({"status": False, "message": "Unauthorized"}, status=403)

            total_pt = Payroll.objects.filter(
                employee__company=company,
                pay_run__status="FINALIZED",
                payroll_period__start_date__month=month,
                payroll_period__start_date__year=year
            ).aggregate(total=Sum("professional_tax"))["total"] or 0

            return Response({
                "status": True,
                "records": {"pt": float(total_pt)}
            })

        except Exception as e:
            return Response({
                "status": False,
                "message": "Failed to load PT report",
                "error": str(e)
            }, status=500)



class PFReportView(APIView):
    permission_classes = (IsAuthenticated,)

    def post(self, request):
        try:
            month = request.data.get("month")
            year = request.data.get("year")

            company = get_company_from_request(request)
            if not company:
                return Response({"status": False, "message": "Unauthorized"}, status=403)

            payrolls = Payroll.objects.filter(
                employee__company=company,
                pay_run__status="FINALIZED",
                payroll_period__start_date__month=month,
                payroll_period__start_date__year=year
            )

            total_pf = payrolls.aggregate(
                total=Sum("provident_fund")
            )["total"] or 0

            return Response({
                "status": True,
                "records": {
                    "pf": float(total_pf),
                    "employee_pf": float(total_pf),
                    "employer_pf": float(total_pf)
                }
            })

        except Exception as e:
            return Response({
                "status": False,
                "message": "Failed to load PF report",
                "error": str(e)
            }, status=500)


class TDSReportView(APIView):
    permission_classes = (IsAuthenticated,)

    def post(self, request):
        try:
            financial_year = request.data.get("financial_year")

            start_year, end_year = map(int, financial_year.split("-"))

            company = get_company_from_request(request)
            if not company:
                return Response({"status": False, "message": "Unauthorized"}, status=403)

            total_tds = Payroll.objects.filter(
                employee__company=company,
                pay_run__status="FINALIZED",
                payroll_period__start_date__year__gte=start_year,
                payroll_period__start_date__year__lte=end_year
            ).aggregate(total=Sum("income_tax"))["total"] or 0

            return Response({
                "status": True,
                "records": {"tds": float(total_tds)}
            })

        except Exception as e:
            return Response({
                "status": False,
                "message": "Failed to load TDS report",
                "error": str(e)
            }, status=500)



class PFExcelExportView(APIView):
    permission_classes = (IsAuthenticated,)

    def post(self, request):
        try:
            month = request.data.get("month")
            year = request.data.get("year")

            company = get_company_from_request(request)
            if not company:
                return Response({"status": False, "message": "Unauthorized"}, status=403)

            payrolls = Payroll.objects.select_related(
                "employee", "employee__user"
            ).filter(
                employee__company=company,
                pay_run__status="FINALIZED",
                payroll_period__start_date__month=month,
                payroll_period__start_date__year=year
            )

            wb = Workbook()
            ws = wb.active
            ws.title = "PF Report"

            ws.append([
                "Employee Name", "Employee ID",
                "Basic Salary", "Employee PF", "Employer PF", "Total PF"
            ])

            total_pf = 0

            for p in payrolls:
                pf = p.provident_fund or 0
                total_pf += pf * 2
                ws.append([
                    p.employee.user.name,
                    str(p.employee.id),
                    float(p.basic_salary),
                    float(pf),
                    float(pf),
                    float(pf * 2)
                ])

            ws.append([])
            ws.append(["", "TOTAL", "", "", "", float(total_pf)])

            response = HttpResponse(
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            response["Content-Disposition"] = f'attachment; filename="PF_{month}_{year}.xlsx"'
            wb.save(response)
            return response

        except Exception as e:
            return Response({
                "status": False,
                "message": "PF Excel export failed",
                "error": str(e)
            }, status=500)



class PTExcelExportView(APIView):
    permission_classes = (IsAuthenticated,)

    def post(self, request):
        try:
            month = request.data.get("month")
            year = request.data.get("year")

            company = get_company_from_request(request)
            if not company:
                return Response({"status": False, "message": "Unauthorized"}, status=403)

            payrolls = Payroll.objects.select_related(
                "employee", "employee__user"
            ).filter(
                employee__company=company,
                pay_run__status="FINALIZED",
                payroll_period__start_date__month=month,
                payroll_period__start_date__year=year,
                professional_tax__gt=0
            )

            wb = Workbook()
            ws = wb.active
            ws.title = "PT Report"

            ws.append(["Employee Name", "Employee ID", "Professional Tax"])

            total_pt = 0
            for p in payrolls:
                total_pt += p.professional_tax
                ws.append([
                    p.employee.user.name,
                    str(p.employee.id),
                    float(p.professional_tax)
                ])

            ws.append([])
            ws.append(["", "TOTAL", float(total_pt)])

            response = HttpResponse(
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            response["Content-Disposition"] = f'attachment; filename="PT_{month}_{year}.xlsx"'
            wb.save(response)
            return response

        except Exception as e:
            return Response({
                "status": False,
                "message": "PT Excel export failed",
                "error": str(e)
            }, status=500)



class TDSExcelExportView(APIView):
    permission_classes = (IsAuthenticated,)

    def post(self, request):
        try:
            financial_year = request.data.get("financial_year")
            start_year, end_year = map(int, financial_year.split("-"))

            company = get_company_from_request(request)
            if not company:
                return Response({"status": False, "message": "Unauthorized"}, status=403)

            payrolls = Payroll.objects.select_related(
                "employee", "employee__user"
            ).filter(
                employee__company=company,
                pay_run__status="FINALIZED",
                payroll_period__start_date__year__gte=start_year,
                payroll_period__start_date__year__lte=end_year
            )

            wb = Workbook()
            ws = wb.active
            ws.title = "TDS Report"

            ws.append(["Employee Name", "Employee ID", "TDS Amount"])

            total = 0
            for p in payrolls:
                total += p.income_tax
                ws.append([
                    p.employee.user.name,
                    str(p.employee.id),
                    float(p.income_tax)
                ])

            ws.append([])
            ws.append(["", "TOTAL", float(total)])

            response = HttpResponse(
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            response["Content-Disposition"] = f'attachment; filename="TDS_{financial_year}.xlsx"'
            wb.save(response)
            return response

        except Exception as e:
            return Response({
                "status": False,
                "message": "TDS Excel export failed",
                "error": str(e)
            }, status=500)



class StatutoryDashboardView(APIView):
    permission_classes = (IsAuthenticated,)

    def post(self, request):
        try:
            employee = Employee.objects.filter(
                user=request.user,
                deleted_at__isnull=True
            ).first()

            if not employee or not employee.company:
                return Response({"status": False, "message": "Unauthorized"}, status=403)

            if employee.role not in ["HR", "ADMIN"]:
                return Response(
                    {"status": False, "message": "Insufficient permissions"},
                    status=403
                )

            month = int(request.data.get("month"))
            year = int(request.data.get("year"))
            financial_year = request.data.get("financial_year")

            company = employee.company

            # ---------------------------
            # Monthly (LOCKED payrolls)
            # ---------------------------
            monthly_qs = Payroll.objects.filter(
                employee__company=company,
                pay_run__status="FINALIZED",
                payroll_period__start_date__month=month,
                payroll_period__start_date__year=year
            )

            monthly = monthly_qs.aggregate(
                pf=Sum("provident_fund"),
                pt=Sum("professional_tax"),
                tds=Sum("income_tax")
            )

            # PF = employee + employer (mirror)
            monthly_pf = (monthly["pf"] or 0) * 2

            # ---------------------------
            # FY aggregation (NORMALIZED)
            # ---------------------------
            start_year, end_year = map(int, financial_year.split("-"))

            fy_qs = Payroll.objects.filter(
                employee__company=company,
                pay_run__status="FINALIZED",
                payroll_period__start_date__year__gte=start_year,
                payroll_period__start_date__year__lte=end_year
            )

            fy = fy_qs.aggregate(
                pf=Sum("provident_fund"),
                pt=Sum("professional_tax"),
                tds=Sum("income_tax")
            )

            fy_pf = (fy["pf"] or 0) * 2

            return Response({
                "status": True,
                "monthly": {
                    "pf": float(monthly_pf),
                    "pt": float(monthly["pt"] or 0),
                    "tds": float(monthly["tds"] or 0),
                },
                "year_to_date": {
                    "pf": float(fy_pf),
                    "pt": float(fy["pt"] or 0),
                    "tds": float(fy["tds"] or 0),
                }
            })

        except Exception as e:
            return Response({
                "status": False,
                "message": "Failed to load statutory dashboard",
                "error": str(e)
            }, status=500)




class StatutoryChallanListView(APIView):
    permission_classes = (IsAuthenticated,)

    def post(self, request):
        try:
            employee = Employee.objects.filter(
                user=request.user,
                deleted_at__isnull=True
            ).select_related("company").first()

            if not employee or not employee.company:
                return Response(
                    {"status": False, "message": "Unauthorized"},
                    status=403
                )

            today = now().date()

            challans = StatutoryChallan.objects.filter(
                company=employee.company
            ).order_by("-year", "-month")

            grouped = {"PF": [], "PT": [], "TDS": []}

            for c in challans:
                # ---- COMPUTED STATUS (NO SAVE) ----
                computed_status = c.status
                if c.status == "DUE" and c.due_date:
                    if c.due_date < today:
                        computed_status = "OVERDUE"

                due_soon = (
                    computed_status == "DUE"
                    and c.due_date
                    and today <= c.due_date <= today + timedelta(days=3)
                )

                grouped[c.statutory_type].append({
                    "id": str(c.id),
                    "type": c.statutory_type,
                    "month": c.month,
                    "year": c.year,
                    "amount": float(c.amount),
                    "due_date": c.due_date,
                    "status": computed_status,
                    "due_soon": due_soon,
                    "receipt": c.receipt.url if c.receipt else None,
                    "payment_reference": c.challan_number,
                    "paid_on": c.paid_date,
                    "is_locked": c.status == "PAID",  # UI helper
                })

            return Response({
                "status": True,
                "records": grouped
            })

        except Exception as e:
            return Response({
                "status": False,
                "message": "Failed to load statutory challans",
                "error": str(e)
            }, status=500)



class StatutoryChallanMarkPaidView(APIView):
    permission_classes = (IsAuthenticated, IsAdminUser)
    parser_classes = (MultiPartParser, FormParser)

    def post(self, request):
        try:
            challan_id = request.data.get("challan_id")
            challan_number = request.data.get("challan_number")
            receipt = request.FILES.get("receipt")

            if not challan_id or not challan_number or not receipt:
                return Response(
                    {"status": False, "message": "All fields required"},
                    status=400
                )

            challan = StatutoryChallan.objects.get(id=challan_id)

            if challan.status == "PAID":
                return Response(
                    {"status": False, "message": "Already paid"},
                    status=400
                )

            challan.challan_number = challan_number
            challan.receipt = receipt
            challan.paid_date = now().date()
            challan.status = "PAID"
            challan.paid_by = request.user  # ✅ AUDIT
            challan.save()

            return Response({
                "status": True,
                "message": "Challan marked as PAID"
            })

        except StatutoryChallan.DoesNotExist:
            return Response(
                {"status": False, "message": "Not found"},
                status=404
            )



class BankDisbursementExportView(APIView):
    permission_classes = (IsAuthenticated,)

    def post(self, request):
        try:
            pay_run_id = request.data.get("pay_run_id")
            if not pay_run_id:
                return Response(
                    {"status": False, "message": "pay_run_id is required"},
                    status=400
                )

            payrun = PayRun.objects.filter(id=pay_run_id, status="FINALIZED").first()
            if not payrun:
                return Response(
                    {"status": False, "message": "Invalid or unfinalized Pay Run"},
                    status=400
                )

            payrolls = Payroll.objects.select_related(
                "employee",
                "employee__user"
            ).filter(
                pay_run=payrun,
                net_salary__gt=0
            )

            wb = Workbook()
            ws = wb.active
            ws.title = "Bank Disbursement"

            headers = [
                "Employee Name",
                "Employee ID",
                "Bank Name",
                "Account Number",
                "IFSC Code",
                "Net Salary",
                "Narration"
            ]

            ws.append(headers)
            for col in range(1, len(headers) + 1):
                ws.cell(row=1, column=col).font = Font(bold=True)

            total_amount = 0

            for p in payrolls:
                emp = p.employee
                total_amount += p.net_salary

                ws.append([
                    emp.user.name,
                    str(emp.id),
                    emp.bank_name,
                    emp.bank_account_number,
                    emp.ifsc_code,
                    float(p.net_salary),
                    f"Salary {payrun.payroll_period.start_date.strftime('%b %Y')}"
                ])

            # Control total
            ws.append([])
            ws.append(["", "", "", "", "TOTAL", float(total_amount), ""])

            response = HttpResponse(
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            response["Content-Disposition"] = (
                f'attachment; filename="Bank_Disbursement_{payrun.id}.xlsx"'
            )

            wb.save(response)
            return response

        except Exception as e:
            return Response({
                "status": False,
                "message": "Failed to export bank disbursement file",
                "error": str(e)
            }, status=500)



class StatutoryChallanBulkUploadView(APIView):
    permission_classes = (IsAuthenticated,)
    parser_classes = (MultiPartParser, FormParser)

    def post(self, request):
        try:
            files = request.FILES.getlist("receipts")

            for f in files:
                challan_id = request.data.get(f.name)
                challan = StatutoryChallan.objects.filter(id=challan_id).first()

                if challan and challan.status != "PAID":
                    challan.receipt = f
                    challan.status = "PAID"
                    challan.paid_on = now().date()
                    challan.save()

            return Response({
                "status": True,
                "message": "Bulk challan upload completed"
            })

        except Exception as e:
            return Response({
                "status": False,
                "message": "Bulk upload failed",
                "error": str(e)
            }, status=500)

